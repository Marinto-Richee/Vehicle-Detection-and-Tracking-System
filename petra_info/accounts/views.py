from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import JsonResponse
from .models import *
from .forms import CCTVForm
from django.contrib import messages
import cv2
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from io import BytesIO
import json
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, redirect
from django.db.models.functions import TruncDate 
from collections import Counter
import os
from datetime import datetime

current_year = datetime.now().year

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'accounts/login.html', {'form': form, 'current_year': current_year})

def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def home_view(request):
    vehicles = Vehicle.objects.all()

    # Filtering by license plate and date range
    license_plate = request.GET.get('license_plate')
    if license_plate:
        vehicles = vehicles.filter(license_plate__icontains=license_plate)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        vehicles = vehicles.filter(timestamp__gte=start_date)
    if end_date:
        vehicles = vehicles.filter(timestamp__lte=end_date)

    # Hourly vehicle activity
    hourly_data = vehicles.extra(select={'hour': 'EXTRACT(HOUR FROM timestamp)'}).values('hour').annotate(count=Count('id'))
    hourly_labels = [f'{i}:00' for i in range(24)]
    hourly_counts = [0] * 24
    for data in hourly_data:
        hourly_counts[int(data['hour'])] = data['count']  # Ensure the index is an integer

    # Vehicle detection by day of the week
    day_of_week_data = vehicles.extra(select={'day_of_week': 'EXTRACT(DOW FROM timestamp)'}).values('day_of_week').annotate(count=Count('id'))
    day_of_week_labels = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    day_of_week_counts = [0] * 7
    for data in day_of_week_data:
        day_of_week_counts[int(data['day_of_week'])] = data['count']

    # Cumulative vehicle count over time
    cumulative_data = vehicles.annotate(date_only=TruncDate('timestamp')).values('date_only').annotate(count=Count('id'))
    cumulative_labels = [entry['date_only'].strftime('%Y-%m-%d') for entry in cumulative_data]
    cumulative_counts = list(Counter({entry['date_only']: entry['count'] for entry in cumulative_data}).values())

    # Completed vs Pending vehicles logic
    completed_vehicles = 0
    pending_vehicles = 0

    # We group by license_plate and ensure that each vehicle has both "In" and "Out"
    for license_plate in vehicles.values('license_plate').distinct():
        in_entries = vehicles.filter(license_plate=license_plate['license_plate'], camera='In').count()
        out_entries = vehicles.filter(license_plate=license_plate['license_plate'], camera='Out').count()

        if in_entries > 0 and out_entries > 0:
            completed_vehicles += 1
        elif in_entries > out_entries:
            pending_vehicles += 1

    # Data for Completed vs Pending chart
    completed_pending_data = [completed_vehicles, pending_vehicles]

    context = {
        'vehicles': vehicles,
        'hourly_labels': json.dumps(hourly_labels),
        'hourly_data': json.dumps(hourly_counts),
        'day_of_week_labels': json.dumps(day_of_week_labels),
        'day_of_week_data': json.dumps(day_of_week_counts),
        'cumulative_labels': json.dumps(cumulative_labels),
        'cumulative_data': json.dumps(cumulative_counts),
        'completed_pending_data': json.dumps(completed_pending_data),  # New data for chart
        'current_year': current_year,
    }

    return render(request, 'accounts/home.html', context)

@login_required
def export_to_excel(request):
    vehicles = Vehicle.objects.all()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    # Create a workbook and worksheet
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Vehicle_Record.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vehicles'

    # Write the header row
    columns = ['ID', 'License Plate', 'Timestamp', 'Category', 'Vehicle Image', 'License Plate Image']
    ws.append(columns)

    # Write the data rows
    for i,vehicle in enumerate(vehicles):
        row = [i, vehicle.license_plate, vehicle.timestamp.strftime('%Y-%m-%d %H:%M:%S'), vehicle.camera]
        ws.append(row)

        # Add the vehicle image to the corresponding row
        if vehicle.image:
            image_path = vehicle.image.path  # Local path of the vehicle image
            img = PILImage.open(image_path)
            img_io = BytesIO()
            img.save(img_io, format='PNG')
            img_io.seek(0)

            img_excel = OpenpyxlImage(img_io)
            img_excel.width = 100  # Resize if needed
            img_excel.height = 100
            ws.add_image(img_excel, f'E{ws.max_row}')  # Insert at the right column and row

        # Add the license plate image to the corresponding row
        if vehicle.license_plate_image:
            license_plate_image_path = vehicle.license_plate_image.path  # Local path of the license plate image
            license_img = PILImage.open(license_plate_image_path)
            license_img_io = BytesIO()
            license_img.save(license_img_io, format='PNG')
            license_img_io.seek(0)

            license_img_excel = OpenpyxlImage(license_img_io)
            license_img_excel.width = 100  # Resize if needed
            license_img_excel.height = 100
            ws.add_image(license_img_excel, f'F{ws.max_row}')  # Insert at the right column and row

    # Save the workbook
    wb.save(response)
    return response

@login_required
def config_cameras(request):
    cameras = CCTV.objects.all()
    statuses = ScriptStatus.objects.all()

    if request.method == 'POST':
        form = CCTVForm(request.POST)
        if form.is_valid() and cameras.count() < 2:  # Allow adding a camera only if less than 2 exist
            form.save()
            return redirect('config_cameras')
    else:
        form = CCTVForm()

    return render(request, 'accounts/config_cameras.html', {'cameras': cameras, 'form': form, 'current_year': current_year, "statuses":statuses})

@login_required
def delete_camera(request, camera_id):
    camera = get_object_or_404(CCTV, id=camera_id)
    camera.delete()
    return redirect('config_cameras')

@login_required
def add_camera(request):
    if request.method == 'POST':
        form = CCTVForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Camera added successfully!')
            return redirect('config_cameras')
    else:
        form = CCTVForm()
    
    return render(request, 'accounts/add_camera.html', {'form': form, 'current_year': current_year})

@login_required
def update_camera(request, camera_id):
    camera = get_object_or_404(CCTV, id=camera_id)

    if request.method == 'POST':
        rtsp_url = request.POST.get('rtsp')
        camera.rtsp = rtsp_url
        camera.save()
        return redirect('config_cameras')

    return redirect('config_cameras')

@login_required
def configure_zones(request, camera_id):
    camera = get_object_or_404(CCTV, id=camera_id)
    rtsp_url = camera.rtsp
    # Capture one frame from the RTSP stream
    cap = cv2.VideoCapture(rtsp_url)
    ret, frame = cap.read()
    # resizes the frame to a smaller size for faster processing
    cap.release()
    if ret:
        # Save the frame as a temporary image file in the MEDIA_ROOT
        frame_path = os.path.join(settings.MEDIA_ROOT, 'temp_frame.jpg')
        cv2.imwrite(frame_path, frame)
        frame_url = os.path.join(settings.MEDIA_URL, 'temp_frame.jpg')  # Create URL to access the image
    else:
        frame_url = None

    return render(request, 'accounts/configure_zones.html', {'camera': camera, 'frame_url': frame_url, 'current_year': current_year})

@login_required
def toggle_detection(request, camera_id):
    camera = get_object_or_404(CCTV, id=camera_id)

    if request.method == 'POST':
        # Toggle the detection status
        start_detection = request.POST.get('start_detection') == 'on'
        print(start_detection)
        camera.start_detection = start_detection
        camera.save()
        messages.success(request, f'Detection {"started" if start_detection else "stopped"} for {camera.name}')

    return redirect('config_cameras')

@login_required
@csrf_exempt 
def save_zones(request, camera_id):
    if request.method == 'POST':
            
        camera = get_object_or_404(CCTV, id=camera_id)
        rtsp_url = camera.rtsp
        # Capture one frame from the RTSP stream
        cap = cv2.VideoCapture(rtsp_url)
        ret, frame = cap.read()

        cap.release()
        data = json.loads(request.body)
        mode = data.get('mode')
        points = data.get('points')
        
        # Find the camera based on the provided camera_id
        try:
            camera = CCTV.objects.get(id=camera_id)
        except CCTV.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Camera not found'}, status=404)

        # Save or update the points based on the selected mode
        if mode == 'polygon':
            # Use update_or_create to overwrite existing PolygonZone entry if it exists
            PolygonZone.objects.update_or_create(
                camera=camera,  # Lookup field
                defaults={'coordinates': points, 'name': mode}  # Fields to update
            )
        elif mode == 'line':
            # Use update_or_create to overwrite existing LineZone entry if it exists
            LineZone.objects.update_or_create(
                camera=camera,  # Lookup field
                defaults={'coordinates': points, 'name': mode}  # Fields to update
            )
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid mode'}, status=400)

        return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'error'}, status=400)